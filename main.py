import sys
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QBrush, QColor, QTextCharFormat
from PySide6.QtWidgets import QApplication, QWidget, QVBoxLayout, QCalendarWidget, QPushButton, QLabel, QFormLayout, \
    QLineEdit, QComboBox, QGridLayout
import pandas as pd
from openpyxl import Workbook, load_workbook
from subprocess import call


class CustomCalendarWidget(QCalendarWidget):
    def __init__(self, mood_data, parent=None):
        super().__init__(parent)
        self.mood_data = mood_data

    def update_background_colors(self):
        # This method updates the background colors of the calendar based on the mood data
        for date in self.mood_data.index:
            selected_date = QDate.fromString(date, Qt.ISODate)
            mood_entry = self.mood_data.loc[date]
            mood_color_map = {
                "Happy": "green",
                "Neutral": "yellow",
                "Sad": "blue",
                "Angry": "red",
                "Excited": "orange",
                "Relaxed": "purple",
            }
            mood_color = mood_color_map.get(mood_entry["Mood"], "white")
            mood_color_with_alpha = QColor(mood_color)
            mood_color_with_alpha.setAlpha(120)

            format = QTextCharFormat()
            format.setBackground(QBrush(mood_color_with_alpha))
            self.setDateTextFormat(selected_date, format)

    def paintCell(self, painter, rect, date):
        super().paintCell(painter, rect, date)

        selected_date = date.toString(Qt.ISODate)

        if date == self.selectedDate():
            selected_color = QColor("gray")
            selected_color.setAlpha(150)
            painter.fillRect(rect, QBrush(QColor(selected_color)))
        elif selected_date in self.mood_data.index:
            mood_entry = self.mood_data.loc[selected_date]
            mood_color_map = {
                "Happy": "green",
                "Neutral": "yellow",
                "Sad": "blue",
                "Angry": "red",
                "Excited": "orange",
                "Relaxed": "purple",
            }

            mood_color = mood_color_map.get(mood_entry["Mood"], "white")

            mood_color_with_alpha = QColor(mood_color)
            mood_color_with_alpha.setAlpha(120)

            painter.fillRect(rect, QBrush(mood_color_with_alpha))
        else:
            self.setStyleSheet("")


class MoodTrackerApp(QWidget):
    def __init__(self):
        super().__init__()
        self.percentage_window = None
        self.init_ui()

    def refresh_calendar(self):
        self.calendar.update_background_colors()
        self.calendar.repaint()

    def init_ui(self):
        self.setWindowTitle("Mood Tracker")
        self.setGeometry(100, 100, 800, 600)

        self.calendar = CustomCalendarWidget(self.load_mood_data(), self)
        self.calendar.clicked.connect(self.show_mood_for_date)

        self.mood_label = QLabel("Mood:")
        self.mood_combobox = QComboBox()
        self.mood_combobox.addItems(["Happy", "Neutral", "Sad", "Angry", "Excited", "Relaxed"])

        self.headache_checkbox = QComboBox()
        self.no_yes_options(self.headache_checkbox)

        self.eat_well_checkbox = QComboBox()
        self.yes_no_options(self.eat_well_checkbox)

        self.sleep_well_checkbox = QComboBox()
        self.yes_no_options(self.sleep_well_checkbox)

        self.stressful_day_checkbox = QComboBox()
        self.no_yes_options(self.stressful_day_checkbox)

        self.medicine_checkbox = QComboBox()
        self.no_yes_options(self.medicine_checkbox)

        self.description_label = QLabel("Description:")
        self.description_edit = QLineEdit()

        self.save_button = QPushButton("Save")
        self.save_button.clicked.connect(self.save_mood_entry)

        self.show_percentage_button = QPushButton("Show Percentage")
        self.show_percentage_button.clicked.connect(lambda: self.show_percentage())

        self.clear_button = QPushButton("Clear Entry")
        self.clear_button.clicked.connect(self.clear_mood_entry)

        self.open_excel_button = QPushButton("Open Excel File")
        self.open_excel_button.clicked.connect(self.open_excel_file)

        # Layout
        layout = QGridLayout(self)
        layout.addWidget(self.calendar, 0, 0, 1, 2)
        layout.setColumnStretch(1, 1)

        form_layout = QFormLayout()
        form_layout.addRow(self.mood_label, self.mood_combobox)
        form_layout.addRow("Headache:", self.headache_checkbox)
        form_layout.addRow("Eat well:", self.eat_well_checkbox)
        form_layout.addRow("Sleep well:", self.sleep_well_checkbox)
        form_layout.addRow("Stressful day:", self.stressful_day_checkbox)
        form_layout.addRow("Medicine:", self.medicine_checkbox)
        form_layout.addRow(self.description_label, self.description_edit)
        form_layout.addRow(self.save_button)
        form_layout.addRow(self.clear_button)
        form_layout.addRow("Show Percentage:", self.show_percentage_button)
        form_layout.addRow("Open Excel File:", self.open_excel_button)

        layout.addLayout(form_layout, 0, 2)

    def reorganize_excel_by_date(self):
        mood_data = self.load_mood_data()

        if not mood_data.empty:
            mood_data.sort_index(inplace=True)

            # Save the sorted DataFrame back to the Excel file
            self.save_mood_data(mood_data)

            # Update the calendar widget with the reorganized data
            self.calendar.mood_data = mood_data
            self.calendar.update_background_colors()

            print("Excel file reorganized by date.")
        else:
            print("No data to reorganize in the Excel file.")

    def yes_no_options(self, combobox):
        combobox.addItems(["Yes", "No"])

    def no_yes_options(self, combobox):
        combobox.addItems(["No", "Yes"])

    def show_mood_for_date(self):
        selected_date = self.calendar.selectedDate().toString(Qt.ISODate)
        mood_data = self.load_mood_data()

        if selected_date in mood_data.index:
            mood_entry = mood_data.loc[selected_date]
            self.mood_combobox.setCurrentText(mood_entry["Mood"])

            self.headache_checkbox.setCurrentIndex(self.headache_checkbox.findText(mood_entry["Headache"]))
            self.eat_well_checkbox.setCurrentIndex(self.eat_well_checkbox.findText(mood_entry["Eat Well"]))
            self.sleep_well_checkbox.setCurrentIndex(self.sleep_well_checkbox.findText(mood_entry["Sleep Well"]))
            self.stressful_day_checkbox.setCurrentIndex(
                self.stressful_day_checkbox.findText(mood_entry["Stressful Day"]))
            self.medicine_checkbox.setCurrentIndex(self.medicine_checkbox.findText(mood_entry["Medicine"]))

            description_text = str(mood_entry["Description"])
            self.description_edit.setText(description_text)

            if selected_date in mood_data.index:
                mood_entry = mood_data.loc[selected_date]
                mood_color_map = {
                    "Happy": "green",
                    "Neutral": "yellow",
                    "Sad": "blue",
                    "Angry": "red",
                    "Excited": "orange",
                    "Relaxed": "purple",
                }

                mood_color = mood_color_map.get(mood_entry["Mood"], "white")
            else:
                self.calendar.setStyleSheet("")

        else:
            self.mood_combobox.setCurrentIndex(0)
            self.headache_checkbox.setCurrentIndex(0)
            self.eat_well_checkbox.setCurrentIndex(0)
            self.sleep_well_checkbox.setCurrentIndex(0)
            self.stressful_day_checkbox.setCurrentIndex(0)
            self.medicine_checkbox.setCurrentIndex(0)
            self.description_edit.clear()
            self.calendar.setStyleSheet("")

    def save_mood_entry(self):
        selected_date = self.calendar.selectedDate().toString(Qt.ISODate)
        mood_data = self.load_mood_data()

        mood_data.loc[selected_date] = {
            "Mood": self.mood_combobox.currentText(),
            "Headache": self.headache_checkbox.currentText(),
            "Eat Well": self.eat_well_checkbox.currentText(),
            "Sleep Well": self.sleep_well_checkbox.currentText(),
            "Stressful Day": self.stressful_day_checkbox.currentText(),
            "Medicine": self.medicine_checkbox.currentText(),
            "Description": self.description_edit.text(),
        }

        self.save_mood_data(mood_data)
        self.calendar.mood_data = mood_data
        self.calendar.update_background_colors()

    def load_mood_data(self):
        try:
            workbook = load_workbook("mood_data.xlsx")
            if "Sheet1" in workbook.sheetnames:
                mood_data = pd.read_excel("mood_data.xlsx", index_col=0)
            else:
                mood_data = pd.DataFrame(
                    columns=["Mood", "Headache", "Eat Well", "Sleep Well", "Stressful Day", "Medicine", "Description"])
        except FileNotFoundError:
            workbook = Workbook()
            mood_data = pd.DataFrame(
                columns=["Mood", "Headache", "Eat Well", "Sleep Well", "Stressful Day", "Medicine", "Description"])

        except Exception as e:
            print(f"Error: {e}")

        workbook.save("mood_data.xlsx")

        return mood_data

    def save_mood_data(self, mood_data):
        mood_data.to_excel("mood_data.xlsx")

    def show_percentage(self):
        mood_data = self.calendar.mood_data

        total_entries = len(mood_data)
        if total_entries == 0:
            print("No entries available for percentage calculation.")
            return

        mood_columns = ["Mood", "Headache", "Eat Well", "Sleep Well", "Stressful Day", "Medicine"]

        mood_percentages = {}
        for column in mood_columns:
            if column in mood_data.columns:
                column_counts = mood_data[column].value_counts().sort_index()
                column_percentages = {value: (count / total_entries) * 100 for value, count in column_counts.items()}
                mood_percentages[column] = column_percentages

        mood_percentage_text = "Percentage of each mood entry:\n"
        for column, percentages in mood_percentages.items():
            mood_percentage_text += f"{column}:\n"
            for value, percentage in percentages.items():
                mood_percentage_text += f"  {value}: {percentage:.2f}%\n"

        if not self.percentage_window:
            self.percentage_window = QWidget()
            self.percentage_window.setWindowTitle("Mood Percentage")

        mood_label = QLabel(mood_percentage_text)

        layout = QVBoxLayout()
        layout.addWidget(mood_label)
        self.percentage_window.setLayout(layout)

        self.percentage_window.show()

    def clear_mood_entry(self):
        selected_date = self.calendar.selectedDate().toString(Qt.ISODate)
        mood_data = self.load_mood_data()

        if selected_date in mood_data.index:
            mood_data = mood_data.drop(index=selected_date)
            self.save_mood_data(mood_data)
            self.calendar.mood_data = mood_data

            # Clear the format for the selected date
            selected_date = QDate.fromString(selected_date, Qt.ISODate)
            format = QTextCharFormat()
            self.calendar.setDateTextFormat(selected_date, format)

            self.calendar.repaint()  # Trigger a repaint to reflect the changes
            self.calendar.setStyleSheet("")  # Clear the style for the selected date
        else:
            print("No entry to clear for the selected date.")

    def open_excel_file(self):
        try:
            call(["start", "mood_data.xlsx"], shell=True)
        except Exception as e:
            print(f"Error opening Excel file: {e}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MoodTrackerApp()
    window.show()
    sys.exit(app.exec())
    
